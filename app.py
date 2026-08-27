"""
SAT — Shift Arrangement Tool

A simple Streamlit web app that:
  1. Lets a user define a date range, colleague names, public holidays, and
     unavailability records.
  2. Auto-generates an even, constraint-respecting duty schedule
     (Morning Health Check, Deployment, Weekend Morning Health Check).
  3. Exports the schedule to a styled .xlsx file (openpyxl).

Run locally:
    streamlit run app.py
"""

import io
from collections import defaultdict
from datetime import date, timedelta

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DEFAULT_NAMES = "Andy, Jessica, Tina, Alan"

WEEKDAY_MORNING_COL = "Morning Health Check (08:45~)"
WEEKEND_MORNING_COL = "Weekend Morning health check (08:45~)"
DEPLOYMENT_COL = "Deployment"
NOT_AVAILABLE_COL = "Not available"
DATE_COL = "Date"

# ---- Chinese weekday helper ------------------------------------------------
CHINESE_WEEKDAYS = {
    0: "星期一",
    1: "星期二",
    2: "星期三",
    3: "星期四",
    4: "星期五",
    5: "星期六",
    6: "星期日",
}


# ---------------------------------------------------------------------------
# Input parsing helpers
# ---------------------------------------------------------------------------


def parse_names(raw: str) -> list[str]:
    """Parse comma/newline separated names into a clean non-empty list."""
    parts = [p.strip() for p in raw.replace("\n", ",").split(",")]
    return [p for p in parts if p]


def parse_holidays(raw: str) -> set[date]:
    """Parse comma-separated YYYY-MM-DD public holidays into a set of dates."""
    holidays: set[date] = set()
    for token in raw.replace("\n", ",").split(","):
        token = token.strip()
        if not token:
            continue
        try:
            holidays.add(date.fromisoformat(token))
        except ValueError:
            continue  # ignore malformed tokens rather than crash
    return holidays


def is_weekend(day: date) -> bool:
    return day.weekday() >= 5  # Sat=5, Sun=6


def needs_weekday_shift(day: date) -> bool:
    """Weekday Morning + Deployment run Mon-Fri excluding public holidays."""
    return not is_weekend(day)


def needs_weekend_shift(day: date, holidays: set[date]) -> bool:
    """Weekend Support runs Sat, Sun and public holidays."""
    return is_weekend(day) or day in holidays


def date_label(day: date) -> str:
    """Format date as e.g. '星期四 2026 07 23'."""
    weekday = CHINESE_WEEKDAYS[day.weekday()]
    return f"{weekday} {day.year} {day.month:02d} {day.day:02d}"


# ---------------------------------------------------------------------------
# Main scheduling engine
# ---------------------------------------------------------------------------

# UNAVAILABLE shape throughout the engine and Excel export:
#   { name: { date: [note, note, ...] } }
# A colleague is considered unavailable on (name, date) regardless of note text.


def build_schedule(
    start: date,
    end: date,
    names: list[str],
    holidays: set[date],
    unavailable: dict[str, dict[date, list[str]]],
) -> tuple[list[dict], dict[str, dict[str, int]]]:
    """
    Generate the schedule row by row.

    Returns ``(rows, counts)``:
      - ``rows`` is a list of dicts ready for preview/Excel.
      - ``counts`` maps name -> {shift_col -> count} for the summary table.
    """
    if not names:
        return [], {}

    # Collect every date in range.
    days: list[date] = []
    cur = start
    while cur <= end:
        days.append(cur)
        cur += timedelta(days=1)

    # Per-colleague per-shift tally (drives even distribution).
    counts: dict[str, dict[str, int]] = {
        n: {WEEKDAY_MORNING_COL: 0, DEPLOYMENT_COL: 0, WEEKEND_MORNING_COL: 0}
        for n in names
    }

    # Rest-state: a person who did Deployment or Weekend Support on the previous
    # day cannot do Morning Health Check today. Rebuilt each day from the
    # immediately preceding day's assignment, so it never accumulates.
    cannot_morning_today: set[str] = set()

    # Weekend Separation: Saturday and Sunday must be assigned to different people.
    last_weekend_person: str | None = None

    rows: list[dict] = []

    for day in days:
        weekday_shift_needed = needs_weekday_shift(day)
        weekend_shift_needed = needs_weekend_shift(day, holidays)

        # Unavailability notes for this date (displayed regardless of shift).
        notes: list[str] = [
            f"{name} ({note})"
            for name, rec in unavailable.items()
            if day in rec
            for note in rec[day]
        ]

        def available_for(name: str) -> bool:
            return day not in unavailable.get(name, {})

        # ---- Morning Health Check (weekday only) ---------------------------
        morning = ""
        if weekday_shift_needed:
            pool = [
                n for n in names if available_for(n) and n not in cannot_morning_today
            ]
            if pool:
                morning = min(pool, key=lambda n: counts[n][WEEKDAY_MORNING_COL])
                counts[morning][WEEKDAY_MORNING_COL] += 1

        # ---- Deployment (weekday only) -------------------------------------
        deployment = ""
        if weekday_shift_needed:
            pool = [n for n in names if available_for(n)]
            if pool:
                # Minimize overlap: nudge away from whoever did Morning today.
                def deployment_key(n: str):
                    return (counts[n][DEPLOYMENT_COL] + (0.5 if n == morning else 0),)

                deployment = min(pool, key=deployment_key)
                counts[deployment][DEPLOYMENT_COL] += 1

        # ---- Weekend Morning Health Check ----------------------------------
        weekend = ""
        if weekend_shift_needed:
            pool = [n for n in names if available_for(n)]
            if pool:
                # Weekend Separation: avoid the previous weekend's person.
                def weekend_key(n: str):
                    same_weekend = 1 if n == last_weekend_person else 0
                    return (same_weekend, counts[n][WEEKEND_MORNING_COL])

                weekend = min(pool, key=weekend_key)
                counts[weekend][WEEKEND_MORNING_COL] += 1
                if is_weekend(day):
                    last_weekend_person = weekend

        # ---- Rest rule bookkeeping -----------------------------------------
        # Deployment / Weekend Support today forbid Morning Health Check only
        # on the very next day's morning. Build the block-set fresh each day so
        # it never accumulates: it holds exactly today's deployment/weekend
        # people, consulted when allocating tomorrow's morning.
        cannot_morning_today = {n for n in (deployment, weekend) if n}

        rows.append(
            {
                DATE_COL: day,
                WEEKDAY_MORNING_COL: morning,
                DEPLOYMENT_COL: deployment,
                WEEKEND_MORNING_COL: weekend,
                NOT_AVAILABLE_COL: ", ".join(notes),
            }
        )

    return rows, counts


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_excel(
    rows: list[dict],
    counts: dict[str, dict[str, int]],
    names: list[str],
    holidays: set[date],
    unavailable: dict[str, dict[date, list[str]]],
) -> bytes:
    """Build the styled .xlsx workbook and return it as bytes.

    Layout:
      - Summary table (rows 1..N+1): header + one row per colleague.
      - Main schedule table starting below the summary.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- Palette ------------------------------------------------------------
    gray_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    gray_body = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    light_orange = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    peach_fill = PatternFill(start_color="FBE5D6", end_color="FBE5D6", fill_type="solid")

    # ---- Summary table ------------------------------------------------------
    summary_headers = [
        "Colleague Name",
        "Morning Health Check (08:45~)",
        "Deployment",
        "Weekend Morning health check (08:45~)",
        "Not available",
    ]
    for col_idx, header in enumerate(summary_headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.fill = gray_header
        c.alignment = center
        c.border = border
        c.font = Font(bold=True)

    for i, name in enumerate(names):
        row = 2 + i
        c = counts.get(name, {})
        not_available_total = sum(len(rec) for rec in unavailable.get(name, {}).values())
        values = [
            name,
            c.get(WEEKDAY_MORNING_COL, 0),
            c.get(DEPLOYMENT_COL, 0),
            c.get(WEEKEND_MORNING_COL, 0),
            not_available_total or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = center
            cell.border = border

    # ---- Main schedule table ------------------------------------------------
    main_headers = [
        "Date",
        "Morning Health Check (08:45~)",
        "Deployment",
        "Weekend Morning health check (08:45~)",
        "Not available",
    ]
    header_fills = {
        1: gray_header,
        2: blue_fill,
        3: gray_body,
        4: blue_fill,
        5: light_orange,
    }

    main_header_row = 2 + len(names) + 1  # blank spacer row after summary
    for col_idx, header in enumerate(main_headers, start=1):
        c = ws.cell(row=main_header_row, column=col_idx, value=header)
        c.fill = header_fills[col_idx]
        c.alignment = center
        c.border = border
        c.font = Font(bold=True)

    name_columns = [WEEKDAY_MORNING_COL, DEPLOYMENT_COL, WEEKEND_MORNING_COL]

    r = main_header_row + 1
    for row_dict in rows:
        day = row_dict[DATE_COL]
        weekend = needs_weekend_shift(day, holidays)

        # Date column: bold, centered, peach fill for weekend/holiday.
        dc = ws.cell(row=r, column=1, value=date_label(day))
        dc.alignment = center
        dc.border = border
        dc.font = Font(bold=True)
        if weekend:
            dc.fill = peach_fill

        # Name columns: centered, thin borders, clear fonts.
        for col_idx, col_name in enumerate(name_columns, start=2):
            cell = ws.cell(row=r, column=col_idx, value=row_dict[col_name])
            cell.alignment = center
            cell.border = border

        # Not available column: soft orange fill when present.
        nc = ws.cell(row=r, column=5, value=row_dict[NOT_AVAILABLE_COL])
        nc.alignment = center
        nc.border = border
        if row_dict[NOT_AVAILABLE_COL]:
            nc.fill = light_orange

        r += 1

    # ---- Column widths ------------------------------------------------------
    ws.column_dimensions["A"].width = 24
    for col_letter in ("B", "C", "D", "E"):
        ws.column_dimensions[col_letter].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------


def default_dates() -> tuple[date, date]:
    today = date.today()
    return today, today + timedelta(days=30)


def main() -> None:
    st.set_page_config(page_title="SAT — Shift Arrangement Tool", page_icon="📅", layout="wide")

    st.title("📅 SAT — Shift Arrangement Tool")
    st.caption("Generate an even, constraint-respecting duty schedule and export it to Excel.")

    # Session state for unavailability records: { (name, date): [notes] }
    if "unavailability" not in st.session_state:
        st.session_state.unavailability = {}

    with st.sidebar:
        st.header("Settings")

        date_range = st.date_input(
            "Time Period",
            value=default_dates(),
            min_value=date(2000, 1, 1),
            max_value=date(2100, 12, 31),
        )
        start, end = date_range[0], date_range[1]
        if start > end:
            st.warning("Start date is after end date; swapping them.")
            start, end = end, start

        st.subheader("Colleague Names")
        names_raw = st.text_area(
            "Comma or newline separated",
            value=DEFAULT_NAMES,
            height=100,
        )
        names = parse_names(names_raw)

        st.subheader("Public Holidays")
        holidays_raw = st.text_area(
            "YYYY-MM-DD, comma-separated",
            value="",
            height=90,
            placeholder="2026-10-01, 2026-10-02",
        )
        holidays = parse_holidays(holidays_raw)

        # ---- Unavailability form ------------------------------------------
        st.subheader("Unavailability")
        with st.form("unavailability_form", clear_on_submit=True):
            u_name = st.selectbox("Colleague", names if names else ["—"])
            u_date = st.date_input(
                "Date",
                value=start,
                min_value=date(2000, 1, 1),
                max_value=date(2100, 12, 31),
            )
            u_note = st.text_input("Reason / Note", placeholder="Night, VL, TO, ...")
            submitted = st.form_submit_button("Add Record")

            if submitted:
                if u_name != "—" and u_note.strip():
                    key = (u_name, u_date)
                    notes = st.session_state.unavailability.setdefault(key, [])
                    if u_note.strip() not in notes:
                        notes.append(u_note.strip())
                    st.rerun()
                elif u_name == "—":
                    st.error("Add a colleague first.")

        # ---- List current unavailability with delete ----------------------
        if st.session_state.unavailability:
            st.markdown("**Current records:**")
            for (uname, udate), notes in list(st.session_state.unavailability.items()):
                for note in notes:
                    col1, col2 = st.columns([4, 1])
                    col1.caption(f"{uname} · {udate} · {note}")
                    btn_key = f"del_{uname}_{udate.isoformat()}_{note}"
                    if col2.button("🗑", key=btn_key):
                        notes.remove(note)
                        if not notes:
                            del st.session_state.unavailability[(uname, udate)]
                        st.rerun()

    # Convert session state into { name: { date: [notes] } } for the engine.
    unavailable_by_name: dict[str, dict[date, list[str]]] = defaultdict(dict)
    for (uname, udate), notes in st.session_state.unavailability.items():
        unavailable_by_name[uname][udate] = list(notes)

    if not names:
        st.error("Please enter at least one colleague name.")
        return

    rows, counts = build_schedule(
        start=start,
        end=end,
        names=names,
        holidays=holidays,
        unavailable=unavailable_by_name,
    )

    if not rows:
        st.error("No days in the selected range.")
        return

    # ---- Preview table -----------------------------------------------------
    st.subheader("Preview")
    df = pd.DataFrame(rows)
    df_display = df.copy()
    df_display[DATE_COL] = df[DATE_COL].apply(date_label)
    st.dataframe(df_display, width="stretch", hide_index=True)

    # ---- Summary stats -----------------------------------------------------
    st.subheader("Shift Counts")
    summary_rows = [
        {
            "Colleague": name,
            WEEKDAY_MORNING_COL: counts.get(name, {}).get(WEEKDAY_MORNING_COL, 0),
            DEPLOYMENT_COL: counts.get(name, {}).get(DEPLOYMENT_COL, 0),
            WEEKEND_MORNING_COL: counts.get(name, {}).get(WEEKEND_MORNING_COL, 0),
        }
        for name in names
    ]
    st.dataframe(pd.DataFrame(summary_rows), width="stretch", hide_index=True)

    # ---- Download button ---------------------------------------------------
    file_bytes = export_excel(rows, counts, names, holidays, unavailable_by_name)

    st.subheader("Export")
    st.download_button(
        label="⬇️ Download Excel Schedule",
        data=file_bytes,
        file_name=f"duty_schedule_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        width="stretch",
    )


if __name__ == "__main__":
    main()